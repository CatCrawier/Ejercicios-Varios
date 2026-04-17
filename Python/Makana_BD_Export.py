import csv
from datetime import datetime
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DB_CONFIG = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "",
    "database": "makana_pizza",
}

TABLAS = [
    "categorias",
    "clientes",
    "productos",
    "promociones",
    "pedidos",
    "detalle_pedidos",
]

COLOR_HEADER_BG = "C0392B"
COLOR_HEADER_FG = "FFFFFF"
COLOR_FILA_PAR = "FADBD8"


def crear_bordes():
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def crear_hoja_tabla(libro_trabajo, cursor_bd, tabla):
    cursor_bd.execute(f"SELECT * FROM `{tabla}`")
    filas = cursor_bd.fetchall()
    columnas = [desc[0] for desc in cursor_bd.description]

    hoja = libro_trabajo.create_sheet(title=tabla)

    for col_idx, col_name in enumerate(columnas, start=1):
        celda = hoja.cell(row=1, column=col_idx, value=col_name)
        celda.font = Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=11)
        celda.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = crear_bordes()
    hoja.row_dimensions[1].height = 22

    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, valor in enumerate(fila, start=1):
            celda = hoja.cell(row=row_idx, column=col_idx, value=valor)
            celda.font = Font(name="Arial", size=10)
            celda.alignment = Alignment(vertical="center")
            celda.border = crear_bordes()
            if row_idx % 2 == 0:
                celda.fill = PatternFill("solid", fgColor=COLOR_FILA_PAR)

    if filas:
        ref = f"A1:{get_column_letter(len(columnas))}{len(filas) + 1}"
        tabla_excel = Table(displayName=f"tbl_{tabla}", ref=ref)
        tabla_excel.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium3",
            showRowStripes=True,
            showFirstColumn=False,
            showLastColumn=False,
            showColumnStripes=False,
        )
        hoja.add_table(tabla_excel)

    for col_idx, col_name in enumerate(columnas, start=1):
        lengths = [len(col_name)]
        lengths.extend(len(str(f[col_idx - 1])) for f in filas if f[col_idx - 1] is not None)
        max_len = max(lengths) if lengths else len(col_name)
        hoja.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    return filas, columnas


def crear_hoja_resumen(libro_trabajo, resumen_tablas):
    hoja = libro_trabajo.create_sheet(title="Indice", index=0)
    hoja.sheet_view.showGridLines = False

    hoja.merge_cells("A1:C1")
    hoja["A1"].value = "Makana Pizza - Exportacion de Base de Datos"
    hoja["A1"].font = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_BG)
    hoja["A1"].alignment = Alignment(horizontal="center", vertical="center")
    hoja.row_dimensions[1].height = 30

    hoja.merge_cells("A2:C2")
    hoja["A2"].value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    hoja["A2"].font = Font(name="Arial", italic=True, size=10, color="888888")
    hoja["A2"].alignment = Alignment(horizontal="center")
    hoja.row_dimensions[2].height = 18

    for col, texto in enumerate(["Tabla", "Filas exportadas", "Hoja"], start=1):
        celda = hoja.cell(row=4, column=col, value=texto)
        celda.font = Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=11)
        celda.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = crear_bordes()
    hoja.row_dimensions[4].height = 20

    for i, (tabla, total) in enumerate(resumen_tablas, start=5):
        for col, valor in enumerate([tabla, total, tabla], start=1):
            celda = hoja.cell(row=i, column=col, value=valor)
            celda.font = Font(name="Arial", size=10)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = crear_bordes()
            if i % 2 == 0:
                celda.fill = PatternFill("solid", fgColor=COLOR_FILA_PAR)

    hoja.column_dimensions["A"].width = 22
    hoja.column_dimensions["B"].width = 20
    hoja.column_dimensions["C"].width = 22


def exportar_csv(archivo_csv, datos):
    with open(archivo_csv, "w", newline="", encoding="utf-8-sig") as archivo:
        writer = csv.writer(archivo, quoting=csv.QUOTE_MINIMAL)
        for tabla, filas, columnas in datos:
            writer.writerow([f"### TABLA: {tabla.upper()} ###"])
            writer.writerow(columnas)
            writer.writerows(filas)
            writer.writerow([])


def principal():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo_excel = f"makana_pizza_{timestamp}.xlsx"
    archivo_csv = f"makana_pizza_{timestamp}.csv"

    conn = None
    cursor = None

    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        print("Conexion exitosa a MySQL\n")

        libro_trabajo = Workbook()
        libro_trabajo.remove(libro_trabajo.active)

        resumen_tablas = []
        datos_para_csv = []

        print("-" * 45)
        print(f"{'Tabla':<25} {'Filas':>6}  {'Excel':^6}  {'CSV':^6}")
        print("-" * 45)

        for tabla in TABLAS:
            try:
                filas, columnas = crear_hoja_tabla(libro_trabajo, cursor, tabla)
                datos_para_csv.append((tabla, filas, columnas))
                resumen_tablas.append((tabla, len(filas)))
                print(f"  {tabla:<23} {len(filas):>6}    OK      OK")
            except mysql.connector.Error as error:
                print(f"  {tabla:<23}  ERROR: {error}")

        print("-" * 45)

        crear_hoja_resumen(libro_trabajo, resumen_tablas)
        exportar_csv(archivo_csv, datos_para_csv)
        libro_trabajo.save(archivo_excel)

        print(f"\nExcel generado: {archivo_excel}")
        print(f"CSV generado: {archivo_csv}")
        print("Exportacion completada correctamente.\n")

    except mysql.connector.Error as error:
        print(f"Error de conexion: {error}")

    finally:
        if cursor is not None:
            cursor.close()
        if conn is not None and conn.is_connected():
            conn.close()
            print("Conexion cerrada.")


if __name__ == "__main__":
    principal()
